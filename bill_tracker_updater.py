import os
import json
import requests
import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

API_KEY           = os.getenv("CONGRESS_API_KEY", "")
EXCEL_PATH        = Path("5.26.26 Climate Tracker v2.xlsx")
STATE_FILE        = Path("bill_tracker_state.json")
DAYS_TO_LOOK_BACK = 7
API_BASE          = "https://api.congress.gov/v3"

KEYWORDS = [
    "climate", "energy", "clean air", "clean water", "emissions",
    "renewable", "carbon", "fossil fuel", "electric", "solar", "wind",
    "nuclear", "pipeline", "greenhouse", "EPA", "environmental",
    "petroleum", "natural gas", "geothermal", "hydropower", "water",
]

HOUSE_SHEET  = "House"
SENATE_SHEET = "Senate"

BILL_TYPE_MAP = {
    "HR": "H.R.", "HRES": "H.RES.", "HJRES": "H.J.RES.", "HCONRES": "H.CON.RES.",
    "S": "S.", "SRES": "S.RES.", "SJRES": "S.J.RES.", "SCONRES": "S.CON.RES.",
}

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_ROW     = 5
FIRST_DATA_ROW = 6


def load_state():
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"seen_ids": [], "last_run": None}


def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


def normalize_bill_number(bn):
    return str(bn).strip().replace(" ", "").replace(".", "").upper()


def extract_bill_num(bill_number):
    match = re.search(r'\d+', str(bill_number))
    return int(match.group()) if match else 0


def get_existing_bill_numbers(ws):
    existing = set()
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=ws.max_row):
        val = row[0].value
        if val:
            existing.add(normalize_bill_number(val))
    return existing


def fetch_bills(from_date, congress=119):
    url = f"{API_BASE}/bill/{congress}"
    params = {
        "api_key": API_KEY,
        "fromDateTime": f"{from_date}T00:00:00Z",
        "sort": "updateDate+desc",
        "limit": 250,
        "format": "json",
    }
    bills  = []
    offset = 0
    while True:
        params["offset"] = offset
        try:
            resp = requests.get(url, params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            log.error(f"API request failed: {e}")
            break
        batch = data.get("bills", [])
        if not batch:
            break
        bills.extend(batch)
        pagination = data.get("pagination", {})
        if len(bills) >= pagination.get("count", 0):
            break
        offset += 250
    log.info(f"Fetched {len(bills)} bills from Congress.gov")
    return bills


def fetch_bill_detail(bill_type, bill_number, congress=119):
    url = f"{API_BASE}/bill/{congress}/{bill_type.lower()}/{bill_number}"
    try:
        resp = requests.get(url, params={"api_key": API_KEY, "format": "json"}, timeout=15)
        resp.raise_for_status()
        return resp.json().get("bill", {})
    except requests.RequestException as e:
        log.warning(f"Could not fetch detail for {bill_type}{bill_number}: {e}")
        return {}


def fetch_cosponsors(bill_type, bill_number, congress=119):
    url = f"{API_BASE}/bill/{congress}/{bill_type.lower()}/{bill_number}/cosponsors"
    try:
        resp = requests.get(url, params={"api_key": API_KEY, "format": "json"}, timeout=15)
        resp.raise_for_status()
        return resp.json().get("cosponsors", [])
    except requests.RequestException as e:
        log.warning(f"Could not fetch cosponsors for {bill_type}{bill_number}: {e}")
        return []


def is_climate_energy_bill(bill):
    title       = (bill.get("title") or "").lower()
    policy_area = (bill.get("policyArea", {}) or {}).get("name", "").lower()
    return any(kw in title or kw in policy_area for kw in KEYWORDS)


def format_bill_number(bill_type, number):
    prefix = BILL_TYPE_MAP.get(bill_type.upper(), bill_type + ".")
    return f"{prefix}{number}"


def format_sponsor(sponsor, is_senate):
    if not sponsor:
        return ""
    first    = sponsor.get("firstName", "")
    last     = sponsor.get("lastName", "")
    party    = sponsor.get("party", "")
    state    = sponsor.get("state", "")
    district = sponsor.get("district", "")
    prefix   = "Sen." if is_senate else "Rep."
    suffix   = f"{party}-{state}"
    if district:
        suffix += f"-{district}"
    return f"{prefix} {first} {last} ({suffix})"


def format_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str[:10], "%Y-%m-%d")
    except ValueError:
        return None


def latest_action_text(detail):
    latest = detail.get("latestAction", {})
    text   = latest.get("text", "")
    date   = latest.get("actionDate", "")
    return f"{text} {date}".strip() if text else ""


def count_cosponsors_by_party(cosponsors):
    dem = rep = ind = 0
    for c in cosponsors:
        party = (c.get("party") or "").upper()
        if party == "D":
            dem += 1
        elif party == "R":
            rep += 1
        else:
            ind += 1
    return dem, rep, ind


def determine_sheet(bill_type):
    bt = bill_type.upper()
    if bt in ("S", "SRES", "SJRES", "SCONRES"):
        return SENATE_SHEET
    return HOUSE_SHEET


def insert_bill_rows(ws, bills_to_insert):
    num_rows = len(bills_to_insert)
    if num_rows == 0:
        return

    # Save all existing hyperlinks before inserting rows
    existing_hyperlinks = {}
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=ws.max_row):
        for cell in row:
            if cell.hyperlink:
                existing_hyperlinks[(cell.row, cell.column)] = cell.hyperlink

    ws.insert_rows(FIRST_DATA_ROW, amount=num_rows)

    # Restore hyperlinks shifted down by the insert
    for (row, col), hyperlink in existing_hyperlinks.items():
        ws.cell(row=row + num_rows, column=col).hyperlink = hyperlink

    # Write new bill rows
    for i, bill in enumerate(bills_to_insert):
        row_idx  = FIRST_DATA_ROW + i
        row_data = [
            bill["bill_number"],
            bill["title"],
            bill["intro_date"],
            bill["sponsor"],
            bill["total_cosponsors"],
            bill["dem_cosponsors"],
            bill["rep_cosponsors"],
            bill["ind_cosponsors"],
            bill["status"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Cambria", size=11)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if col_idx == 3 and value is not None:
                cell.number_format = "dddd, d. mmmm yyyy"
        log.info(f"  -> Inserted: {row_data[0]} | {row_data[2]}")


def update_excel(new_bills):
    if not EXCEL_PATH.exists():
        log.error(f"Excel file not found: {EXCEL_PATH}")
        return

    wb        = load_workbook(EXCEL_PATH)
    house_ws  = wb[HOUSE_SHEET]
    senate_ws = wb[SENATE_SHEET]

    existing_house  = get_existing_bill_numbers(house_ws)
    existing_senate = get_existing_bill_numbers(senate_ws)

    house_bills = [
        b for b in new_bills
        if b["sheet"] == HOUSE_SHEET
        and normalize_bill_number(b["bill_number"]) not in existing_house
    ]
    senate_bills = [
        b for b in new_bills
        if b["sheet"] == SENATE_SHEET
        and normalize_bill_number(b["bill_number"]) not in existing_senate
    ]

    house_bills  = sorted(house_bills,  key=lambda x: (x["intro_date"] or datetime.min, extract_bill_num(x["bill_number"])), reverse=True)
    senate_bills = sorted(senate_bills, key=lambda x: (x["intro_date"] or datetime.min, extract_bill_num(x["bill_number"])), reverse=True)

    log.info(f"New House bills to insert: {len(house_bills)}")
    log.info(f"New Senate bills to insert: {len(senate_bills)}")

    insert_bill_rows(house_ws, house_bills)
    insert_bill_rows(senate_ws, senate_bills)

    wb.save(EXCEL_PATH)
    log.info(f"File saved. Size: {EXCEL_PATH.stat().st_size} bytes")


def build_bill_record(bill, from_date):
    bill_type    = bill.get("type", "")
    bill_number  = str(bill.get("number", ""))
    congress     = bill.get("congress", 119)
    is_senate    = determine_sheet(bill_type) == SENATE_SHEET
    detail       = fetch_bill_detail(bill_type, bill_number, congress)
    cosponsors   = fetch_cosponsors(bill_type, bill_number, congress)
    intro_date   = detail.get("introducedDate", "")
    if intro_date and intro_date < from_date:
        log.info(f"Skipping (introduced {intro_date}): {bill_type} {bill_number}")
        return None
    dem, rep, ind = count_cosponsors_by_party(cosponsors)
    sponsor_info  = (detail.get("sponsors") or [{}])[0] if detail.get("sponsors") else {}
    return {
        "sheet":            determine_sheet(bill_type),
        "bill_number":      format_bill_number(bill_type, bill_number),
        "title":            bill.get("title", ""),
        "intro_date":       format_date(intro_date),
        "sponsor":          format_sponsor(sponsor_info, is_senate),
        "total_cosponsors": len(cosponsors),
        "dem_cosponsors":   dem,
        "rep_cosponsors":   rep,
        "ind_cosponsors":   ind,
        "status":           latest_action_text(detail),
        "_uid":             f"{bill_type}{bill_number}-{congress}",
    }


def run():
    log.info("=== Bill Tracker Updater ===")
    log.info(f"API_KEY present: {bool(API_KEY)}")
    log.info(f"Excel file exists: {EXCEL_PATH.exists()}")

    if not API_KEY:
        log.error("No API key found. Make sure CONGRESS_API_KEY is set.")
        return

    state    = load_state()
    seen_ids = set(state.get("seen_ids", []))
    last_run = state.get("last_run")

    if last_run:
        from_date = last_run[:10]
    else:
        from_date = (datetime.today() - timedelta(days=DAYS_TO_LOOK_BACK)).strftime("%Y-%m-%d")

    log.info(f"Searching for bills introduced since {from_date}")

    raw_bills   = fetch_bills(from_date)
    new_records = []

    for bill in raw_bills:
        uid = f"{bill.get('type')}{bill.get('number')}-{bill.get('congress', 119)}"
        if uid in seen_ids:
            continue
        if not is_climate_energy_bill(bill):
            continue

        log.info(f"Checking: {bill.get('type')} {bill.get('number')} - {bill.get('title', '')[:60]}")
        record = build_bill_record(bill, from_date)
        if record:
            new_records.append(record)
            seen_ids.add(uid)

    if new_records:
        log.info(f"Found {len(new_records)} new bill(s). Updating tracker...")
        update_excel(new_records)
    else:
        log.info("No new climate/energy bills found since last run.")

    state["seen_ids"] = list(seen_ids)
    state["last_run"] = datetime.utcnow().isoformat()
    save_state(state)
    log.info("Done.")


if __name__ == "__main__":
    run()
